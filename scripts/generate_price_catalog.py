import json
import re
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\herik\Downloads\PRECOS_CELULAR_v3_2026.xlsx")

KNOWN_BRANDS = {
    "APPLE": "Apple",
    "SAMSUNG": "Samsung",
    "MOTOROLA": "Motorola",
    "XIAOMI": "Xiaomi",
    "REALME": "Realme",
    "LG": "LG",
    "LENOVO": "Lenovo",
    "ASUS": "Asus",
    "TABLET": "Tablet",
}


SERVICE_MAP = [
    ("tela-paralela", "Troca de Tela", "TELA PARALELA", "1ª Linha"),
    ("tela-premium", "Troca de Tela", "TELA PREMIUM", "Tela Premium"),
    ("tela-original", "Troca de Tela", "TELA ORIGINAL", "Sams"),
    ("bateria-paralela", "Troca de Bateria", "BATERIA PARALELA", None),
    ("bateria-premium", "Troca de Bateria", "BATERIA PREMIUM", "Bat Prem"),
    ("bateria-original", "Troca de Bateria", "BATERIA ORIGINAL", "Bat Orig"),
    ("tampa", "Troca de Tampa", None, "Tampa"),
    ("conector-carga", "Troca de Conector de Carga", None, "Con. Carga"),
    ("carcaca", "Troca de Carcaça", None, "Carcaça"),
]


def normalize(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def parse_money(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("R$", "").strip()
    matches = re.findall(r"\d+(?:[.,]\d+)?", text)
    if not matches:
        return None

    values = []
    for match in matches:
        try:
            values.append(float(match.replace(",", ".")))
        except ValueError:
            continue

    return max(values) if values else None


def build_catalog(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = next((sheet for sheet in wb.worksheets if "Tabela Completa" in sheet.title), wb.active)
    headers = {str(ws.cell(3, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(3, col).value}
    if "Modelo" not in headers:
        headers = {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}
    catalog = []
    seen = {}
    current_brand = "Outros"

    for row in range(2, ws.max_row + 1):
        model = ws.cell(row, headers["Modelo"]).value
        if not model:
            continue

        model = str(model).strip()
        brand_name = KNOWN_BRANDS.get(model.upper())
        if brand_name:
            current_brand = brand_name
            continue

        brand = current_brand
        base_id = normalize(f"{brand}-{model}")
        seen[base_id] = seen.get(base_id, 0) + 1
        item_id = base_id if seen[base_id] == 1 else f"{base_id}-{seen[base_id]}"

        tampa_price = parse_money(ws.cell(row, headers.get("Tampa", 0)).value)
        tampa2_price = parse_money(ws.cell(row, headers.get("Tampa 2", 0)).value)

        services = []
        for key, label, quality, source in SERVICE_MAP:
            value = None
            source_label = source or "Cadastro manual"
            if source == "Tampa":
                value = tampa_price if tampa_price is not None else tampa2_price
                if tampa_price is None and tampa2_price is not None:
                    source_label = "Tampa 2"
            elif source and source in headers:
                value = parse_money(ws.cell(row, headers[source]).value)

            if value is None:
                continue

            services.append({
                "key": key,
                "label": label,
                "sourceLabel": source_label,
                "quality": quality,
                "finalPrice": value,
                "installmentPrice": None,
                "costPrice": None,
                "note": None,
            })

        if any(service["finalPrice"] is not None for service in services):
            catalog.append({
                "id": item_id,
                "brand": brand,
                "model": model,
                "search": f"{brand} {model}".lower(),
                "services": services,
            })

    return catalog


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    output = ROOT / "src" / "data" / "priceCatalog.ts"
    catalog = build_catalog(xlsx_path)
    content = "/* Auto-generated from PRECOS CELULAR v2 2026.xlsx. */\n"
    content += "import type { PriceCatalogItem } from '@/types/pricing'\n\n"
    content += "export const PRICE_CATALOG: PriceCatalogItem[] = "
    content += json.dumps(catalog, ensure_ascii=False, indent=2)
    content += "\n"
    output.write_text(content, encoding="utf-8")
    print(f"Generated {len(catalog)} catalog items at {output}")


if __name__ == "__main__":
    main()
